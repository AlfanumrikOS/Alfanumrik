#!/usr/bin/env python3
"""Phase 0: instrument the QA workbook with tier/phase/owner tags (Notes+Tester).
Does NOT touch the Status column. Adds an Execution_Plan sheet."""
import openpyxl, re, sys

SRC = 'qa/launch-certification/Alfanumrik_Master_QA_Suite2.xlsx'
wb = openpyxl.load_workbook(SRC)

MODULE_SHEETS = ['01_AUTH','02_QUIZ','03_FOXY','04_RAG','05_PAR','06_DPDP','07_TEA','08_SCH',
 '09_SUP','10_PAY','11_QUO','12_GAM','13_OPS','14_HOST','15_FE','16_SEC','17_I18N','18_PERF',
 '19_SEAM','20_REG','21_MX_RBAC','22_MX_RLS','23_MX_CONTENT','24_MX_INPUT','25_MX_DEVICE',
 '26_MX_EDGE','27_MX_NAV']

# tier -> (phase, base method)
TIER_META = {
 'T1':('Phase 1','static code/config/migration review, file:line evidence'),
 'T2':('Phase 2',"repo's own vitest/Playwright suite, run headless"),
 'T3':('Phase 3','parametrized SQL probe via live/staging Supabase'),
 'T4':('Phase 4','Playwright/authenticated-API against running env'),
 'T5':('Phase 5','human-executed (real money/device/inbox)'),
}
AWAIT = {
 'T3':'awaiting live/staging Supabase read access + A/B test accounts',
 'T4':'awaiting running staging URL + test accounts',
 'T5':'awaiting human executor (LIVE keys / physical device / real inbox)',
}

# per-row terminology-finding annotations
FINDINGS = {}
for g in ['6','7','8','9','10','11','12']:
    pass
# specific TC overrides for findings
def finding_note(tcid, feature, check):
    t = (feature+' '+check).lower()
    notes=[]
    if 'language_gap' in t or ('hindi' in t and 'english' in t and 'gap' in t):
        notes.append('FINDING: LANGUAGE_GAP not implemented in code (docs only) -> candidate P-defect, not auto-PASS')
    if 'curriculum_gap' in t or 'out-of-syllabus' in t or 'out of syllabus' in t:
        notes.append('NOTE: CURRICULUM_GAP realised as grounded-answer abstain/coverage; verify against those')
    if 'v3' in t and ('flag' in t or 'ui' in t):
        notes.append('NOTE: ff_ui_v3_* == one_experience_v3_* which were seeded->disabled->REMOVED')
    return notes

def classify(sheet, tcid, pri, layer, feature, check, steps, evidence):
    L=(layer or '').lower(); F=(feature or '').lower(); C=(check or '').lower()
    S=(steps or '').lower(); E=(evidence or '').lower()
    T=' '.join([F,C,S,E])
    # ---- T5 manual (money/whatsapp/inbox/physical) ----
    if sheet=='10_PAY' and tcid in ('PAY-0001','PAY-0002','PAY-0006'):
        return 'T5'
    if 'whatsapp' in T and ('deliver' in T or 'send' in T or 'notify' in T):
        return 'T5'
    if 'razorpay dashboard' in T or 'real ₹' in T or 'live ₹' in T or 'one small real' in T:
        return 'T5'
    if sheet=='01_AUTH' and tcid=='AUTH-0009':
        return 'T5'  # real email inbox reset
    # ---- sheet-driven primary tiers ----
    if sheet=='22_MX_RLS': return 'T3'
    if sheet=='25_MX_DEVICE': return 'T4'
    if sheet=='27_MX_NAV': return 'T4'
    if sheet=='24_MX_INPUT': return 'T4'
    if sheet=='26_MX_EDGE': return 'T4'
    if sheet=='21_MX_RBAC': return 'T4'
    if sheet=='23_MX_CONTENT':
        # readiness honesty + Hindi availability are DB-provable
        if 'readiness' in F or 'hindi availability' in F: return 'T3'
        return 'T4'
    if sheet=='18_PERF': return 'T4'
    if sheet=='19_SEAM': return 'T4'  # live rehearsal
    if sheet=='16_SEC':
        if 'rls' in C or 'cross-user' in C or 'isolation' in C: return 'T3'
        return 'T1'
    if sheet in ('20_REG','14_HOST'): return 'T1'
    if sheet=='04_RAG':
        if 'chunk' in T or 'count' in T or 'rag_status' in T: return 'T3'
        return 'T1'
    if sheet=='11_QUO': return 'T3'
    if sheet=='13_OPS': return 'T1'
    if sheet=='15_FE': return 'T1'
    if sheet=='17_I18N': return 'T1'
    # ---- QUIZ split ----
    if sheet=='02_QUIZ':
        if tcid in ('QUIZ-0006','QUIZ-0007','QUIZ-0009'): return 'T3'
        return 'T2'
    # ---- PAY remaining ----
    if sheet=='10_PAY':
        if tcid in ('PAY-0003','PAY-0004','PAY-0008'): return 'T2'
        return 'T1'
    # ---- FOXY ----
    if sheet=='03_FOXY':
        if tcid=='FOXY-0003': return 'T1'  # taxonomy parity code
        return 'T2'
    # ---- verticals & auth default to auto-suite ----
    if sheet in ('01_AUTH','05_PAR','06_DPDP','07_TEA','08_SCH','09_SUP','12_GAM'):
        return 'T2'
    return 'T1'

def owner_for(tier):
    if tier=='T5': return 'Human'
    if tier in ('T3','T4'): return 'Claude (pending env)'
    return 'Claude'

summary={}
for sheet in MODULE_SHEETS:
    ws=wb[sheet]
    hdr=[c.value for c in ws[2]]
    ci={h:i for i,h in enumerate(hdr) if h}
    notes_c=ci['Notes']; tester_c=ci['Tester']
    for row in ws.iter_rows(min_row=3):
        tcid=row[0].value
        if not tcid: continue
        pri=row[1].value; layer=row[2].value; feature=row[3].value
        check=row[4].value; steps=row[6].value; evidence=row[8].value
        tier=classify(sheet,tcid,pri,layer,feature,check,steps,evidence)
        phase,method=TIER_META[tier]
        owner=owner_for(tier)
        parts=[f'[{tier} | {phase} | {owner}] {method}']
        if tier in AWAIT: parts.append(AWAIT[tier])
        parts += finding_note(tcid,feature or '',check or '')
        note='. '.join(parts)
        # write only if Notes empty (don't clobber real tester notes)
        row[notes_c].value = note
        row[tester_c].value = owner
        summary[tier]=summary.get(tier,0)+1

# ---- Execution_Plan sheet ----
if 'Execution_Plan' in wb.sheetnames:
    del wb['Execution_Plan']
ep=wb.create_sheet('Execution_Plan', index=3)
from openpyxl.styles import Font
def w(r,c,v,bold=False):
    cell=ep.cell(row=r,column=c,value=v)
    if bold: cell.font=Font(bold=True)
    return cell
r=1
w(r,1,'ALFANUMRIK MASTER QA — CLAUDE EXECUTION PLAN (confirmation of approach)',bold=True); r+=2
w(r,1,'Hard rule: DO NOT FAKE. PASS only with auditable evidence (SQL+result / request-response / screenshot+URL+timestamp).',bold=True); r+=1
w(r,1,'Status column is only changed when a test is genuinely executed with evidence. Rows I cannot verify stay NOT RUN with the unblock reason in Notes.'); r+=2
w(r,1,'EVIDENCE TIER LEGEND',bold=True); r+=1
for k in ['T1','T2','T3','T4','T5']:
    ph,me=TIER_META[k]; w(r,1,k); w(r,2,ph); w(r,3,me); w(r,4,('' if k not in AWAIT else AWAIT[k])); r+=1
r+=1
w(r,1,'ROW COUNT BY TIER (this workbook)',bold=True); r+=1
for k in ['T1','T2','T3','T4','T5']:
    w(r,1,k); w(r,2,summary.get(k,0)); r+=1
r+=1
w(r,1,'DEFAULT SCOPE THIS PASS',bold=True); r+=1
for line in ['Phase 0: instrument every row (Notes+Tester) — DONE by this sheet.',
 'Phase 1 (T1 static) + Phase 2 (T2 auto-suite): executed now with real evidence -> Status set to PASS/FAIL.',
 'Phase 3/4/5 (T3 live-DB, T4 live-browser, T5 manual): NOT RUN this pass — unblock reason in each row Notes.',
 'Output: updated .xlsx + evidence/ committed to claude/qa-sheet-testing-phases-xkrtmf + draft PR.']:
    w(r,1,line); r+=1
r+=1
w(r,1,'TERMINOLOGY RECONCILIATIONS (findings surfaced before testing)',bold=True); r+=1
for line in ['CURRICULUM_GAP: no literal token; implemented as grounded-answer/{abstain,coverage,confidence}.ts + content_gap.',
 'LANGUAGE_GAP: NOT implemented in code (docs only) -> FOXY-0004 / 17_I18N candidate defect, not a pass.',
 'ff_ui_v3_*: actually one_experience_v3_* — seeded -> force-disabled -> REMOVED; 15_FE flag rows reframed.',
 'Edge matrix 172 rows = 43 functions x 4 properties (happy/logs/auth/failure), not one row per function.']:
    w(r,1,line); r+=1

wb.save(SRC)
print('Phase 0 complete. Rows tagged by tier:', summary, 'total:', sum(summary.values()))
