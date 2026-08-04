import openpyxl
from datetime import date
SRC='qa/launch-certification/Alfanumrik_Master_QA_Suite2.xlsx'
wb=openpyxl.load_workbook(SRC)
TODAY=date(2026,8,4).isoformat()
EVLOG='evidence/T2_certified_rows_testlog.txt'

# TCID -> (status, tester, defect, evidence/finding note)
PASS={
 ('02_QUIZ','QUIZ-0001'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: api/quiz-server-shuffle-authority.test.ts (server never sends correct key; submit carries only chosen index). Evidence: {EVLOG}.'),
 ('02_QUIZ','QUIZ-0002'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: anti-cheat-server-parity.test.ts + quiz-server-shuffle-authority.test.ts (forged correctness flag ignored; server grades by own key). Evidence: {EVLOG}.'),
 ('10_PAY','PAY-0003'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: api/payments/verify-hmac-reject.test.ts (bad/absent signature rejected before any entitlement). Evidence: {EVLOG}.'),
 ('10_PAY','PAY-0004'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: payments/webhook-retry-and-dedupe-semantics.test.ts + webhook-concurrent-fire.test.ts (replayed/duplicate event id -> exactly one activation, no double-grant). Evidence: {EVLOG}.'),
 ('03_FOXY','FOXY-0001'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: foxy-grounded-gate.test.ts (in-syllabus answer gated on retrieved grounded chunks). Evidence: {EVLOG}.'),
 ('03_FOXY','FOXY-0002'):('PASS','Claude (T2 auto)','', f'PASS via repo suite: api/foxy/structured-abstain-and-history.test.ts + grounded-failure-fallback.test.ts (no-context -> structured abstain, no hallucination). NOTE: realised as grounded-answer abstain/coverage, not a literal CURRICULUM_GAP token. Evidence: {EVLOG}.'),
}
# findings: keep Status NOT RUN, stamp Defect ID + finding note (do not fake PASS/FAIL without live obs)
FIND={
 ('03_FOXY','FOXY-0004'):('QA-FIND-001','FINDING (code gap): explicit Hindi->LANGUAGE_GAP contract NOT implemented (grep LANGUAGE_GAP = 0 hits; docs-only). Candidate P1. Confirm live whether Hindi is answered in Hindi vs silently English before PASS/FAIL. Evidence: evidence/T1_static_findings.txt.'),
 ('16_SEC','SEC-0003'):('QA-FIND-002','FINDING: baseline has SECURITY DEFINER fns but 0 explicit REVOKE-from-anon; positive confirmation that no destructive fn is anon-executable required (architect). Evidence: evidence/T1_static_findings.txt.'),
}

def apply(sheet,tcid,status,tester,defect,note):
    ws=wb[sheet]; hdr=[c.value for c in ws[2]]; ci={h:i for i,h in enumerate(hdr) if h}
    for row in ws.iter_rows(min_row=3):
        if row[0].value==tcid:
            if status is not None: row[ci['Status']].value=status
            if tester is not None: row[ci['Tester']].value=tester
            row[ci['Date']].value=TODAY
            if defect: row[ci['Defect ID']].value=defect
            row[ci['Notes']].value=note
            return True
    return False

n=0
for (sh,tc),(st,te,de,no) in PASS.items():
    if apply(sh,tc,st,te,de,no): n+=1
for (sh,tc),(de,no) in FIND.items():
    if apply(sh,tc,None,None,de,no): n+=1

# also stamp all 17_I18N LANGUAGE_GAP-dependent rows with the finding defect id (keep NOT RUN)
ws=wb['17_I18N']; hdr=[c.value for c in ws[2]]; ci={h:i for i,h in enumerate(hdr) if h}
for row in ws.iter_rows(min_row=3):
    if not row[0].value: continue
    txt=(str(row[3].value)+' '+str(row[4].value)).lower()
    if 'language_gap' in txt or ('hindi' in txt and 'english' in txt):
        if not row[ci['Defect ID']].value:
            row[ci['Defect ID']].value='QA-FIND-001'
        cur=row[ci['Notes']].value or ''
        row[ci['Notes']].value=cur+' | linked QA-FIND-001 (LANGUAGE_GAP not implemented; confirm live).'
        n+=1

wb.save(SRC)
print('applied updates to', n, 'rows')

# recompute Dashboard
wb2=openpyxl.load_workbook(SRC)
mods=[('01_AUTH','01_AUTH'),('02_QUIZ','02_QUIZ'),('03_FOXY','03_FOXY'),('04_RAG','04_RAG'),('05_PAR','05_PAR'),('06_DPDP','06_DPDP'),('07_TEA','07_TEA'),('08_SCH','08_SCH'),('09_SUP','09_SUP'),('10_PAY','10_PAY'),('11_QUO','11_QUO'),('12_GAM','12_GAM'),('13_OPS','13_OPS'),('14_HOST','14_HOST'),('15_FE','15_FE'),('16_SEC','16_SEC'),('17_I18N','17_I18N'),('18_PERF','18_PERF'),('19_SEAM','19_SEAM'),('20_REG','20_REG'),('21_MX_RBAC','21_MX_RBAC'),('22_MX_RLS','22_MX_RLS'),('23_MX_CONTENT','23_MX_CONTENT'),('24_MX_INPUT','24_MX_INPUT'),('25_MX_DEVICE','25_MX_DEVICE'),('26_MX_EDGE','26_MX_EDGE'),('27_MX_NAV','27_MX_NAV')]
tot=pas=fail=nr=0
for _,sh in mods:
    ws=wb2[sh]; hdr=[c.value for c in ws[2]]; ci={h:i for i,h in enumerate(hdr) if h}
    for row in ws.iter_rows(min_row=3):
        if not row[0].value: continue
        s=row[ci['Status']].value; tot+=1
        if s=='PASS': pas+=1
        elif s=='FAIL': fail+=1
        else: nr+=1
print(f'ROLLUP total={tot} pass={pas} fail={fail} notrun/other={nr}')
